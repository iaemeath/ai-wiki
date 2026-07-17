"""Parse .xlsx files into RequirementItem lists using openpyxl."""

from __future__ import annotations

import logging
import os
from typing import Optional

from openpyxl import load_workbook

from app.models.schemas import RequirementItem

logger = logging.getLogger(__name__)


class ExcelParseError(Exception):
    """Raised when the Excel file cannot be parsed."""


def parse_requirements_from_excel(file_path: str) -> list[RequirementItem]:
    """Read an .xlsx file and return a list of RequirementItem.

    Expected sheet structure (first sheet, header row detected automatically):
      - Column A: Requirement ID (e.g. 'REQ-001')
      - Column B: Requirement content / description
    Ignores empty rows and respects a header row that contains common keywords
    (id, req, requirement, 编号, 需求).

    Args:
        file_path: Absolute path to the .xlsx file.

    Returns:
        List of RequirementItem parsed from the sheet.

    Raises:
        ExcelParseError: If the file is missing, invalid, or contains no valid rows.
    """
    if not os.path.isfile(file_path):
        raise ExcelParseError(f"File not found: {file_path}")

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelParseError(f"Failed to open Excel file: {exc}") from exc

    sheet = wb.active
    if sheet is None:
        wb.close()
        raise ExcelParseError("Excel file has no active sheet")

    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ExcelParseError("Excel sheet is empty")

    header_keywords = {"id", "req", "requirement", "编号", "需求", "content", "description", "name"}
    start_row = 0

    # Detect header row
    if rows[0][0] is not None:
        first_cell = str(rows[0][0]).strip().lower()
        if any(kw in first_cell for kw in header_keywords):
            start_row = 1

    items: list[RequirementItem] = []
    seen_ids: set[str] = set()

    for i, row in enumerate(rows[start_row:], start=start_row + 1):
        col_a = row[0]
        col_b = row[1] if len(row) > 1 else None

        # Skip fully empty rows
        if col_a is None and col_b is None:
            continue

        req_id = str(col_a).strip() if col_a is not None else f"REQ-{i:03d}"
        content = str(col_b).strip() if col_b is not None else ""

        if not content:
            logger.debug("Skipping row %d: empty requirement content", i)
            continue

        # Deduplicate IDs
        original_id = req_id
        counter = 1
        while req_id in seen_ids:
            req_id = f"{original_id}-{counter}"
            counter += 1
        seen_ids.add(req_id)

        items.append(
            RequirementItem(
                id=req_id,
                content=content,
                status="pending",
                answers=[],
            )
        )

    if not items:
        raise ExcelParseError(
            "No valid requirement rows found in Excel. "
            "Expected columns: Requirement ID | Description"
        )

    logger.info("Parsed %d requirements from %s", len(items), file_path)
    return items
